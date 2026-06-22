# -*- coding: utf-8 -*-
"""
xlsx2extracted.py —— 把 raw/ 下的 Excel 转成 AI 友好的结构化 Markdown 快照。

每个工作表 = 一个 .md：YAML 头（来源/表头行/行列数）+ 字段字典 + 完整数据表。
与 reference-data/ 既有产出格式一致。纯确定性、可重跑。

表头识别：前 3 行里非空单元格最多的那行即真实列名行（1-based 记入 header_row）。
数据：列名行以下全部非空行；空单元格留空；换行→<br>；datetime→YYYY-MM-DD。
空列（无标题）记为「(空列_X)」。整表无数据行的工作表跳过（记入 SKIPPED）。

用法：
  python tools/xlsx2extracted.py "raw/　51Sim需求管理】.xlsx" "reference-data/需求管理"
"""
import datetime
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

EXTRACT_DATE = datetime.date.today().isoformat()

# 模板/表单类工作表（在线表格自带的空壳，非业务内容）——确定性跳过
SKIP_SHEET_PATTERNS = [r"^表单\d*$", r"^智能表\d*$", r"^-+$"]


def is_skip_sheet(title):
    return any(re.match(p, title.strip()) for p in SKIP_SHEET_PATTERNS)


def nz(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def cell_md(v):
    """单元格转 Markdown：换行→<br>，竖线转义。"""
    s = nz(v).replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("\n", "<br>").replace("|", "\\|")
    return s


def safe_name(s):
    s = str(s).replace("/", "_").replace("\\", "_")
    s = re.sub(r'[:*?"<>|]', "", s).strip()
    return s or "未命名"


def detect_header_row(rows, scan=3):
    """前 scan 行里非空单元格最多者为表头（返回 0-based 索引）。"""
    best_i, best_n = 0, -1
    for i in range(min(scan, len(rows))):
        n = sum(1 for c in rows[i] if nz(c) != "")
        if n > best_n:
            best_i, best_n = i, n
    return best_i


def main():
    if len(sys.argv) < 3:
        print("用法: python tools/xlsx2extracted.py <src.xlsx> <out_dir>")
        sys.exit(1)
    src, out_dir = sys.argv[1], sys.argv[2]
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(src, data_only=True)
    src_name = Path(src).name
    index, skipped = [], []
    seq = 0

    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        # 去掉尾部整行全空
        while rows and all(nz(c) == "" for c in rows[-1]):
            rows.pop()
        if not rows:
            skipped.append((ws.title, "整表为空"))
            continue
        if is_skip_sheet(ws.title):
            skipped.append((ws.title, "模板/表单壳，跳过"))
            continue

        hr = detect_header_row(rows)
        header = rows[hr]
        ncol = len(header)
        # 列名：空标题→(空列_X)
        colnames = []
        for j in range(ncol):
            name = nz(header[j])
            colnames.append(name if name else f"(空列_{get_column_letter(j + 1)})")

        # 数据行：表头之下的非空行
        data = []
        for r in rows[hr + 1:]:
            r = list(r) + [None] * (ncol - len(r))
            if all(nz(c) == "" for c in r[:ncol]):
                continue
            data.append(r[:ncol])

        if not data:
            skipped.append((ws.title, "无数据行"))
            continue

        seq += 1
        fname = f"{seq:02d}-{safe_name(ws.title)}.md"
        title = ws.title

        lines = []
        lines.append("---")
        lines.append(f'source_file: "{src_name}"')
        lines.append(f'sheet: "{title}"')
        lines.append(f"header_row: {hr + 1}")
        lines.append(f"data_rows: {len(data)}")
        lines.append(f"columns: {ncol}")
        lines.append(f"extracted: {EXTRACT_DATE}")
        lines.append("---")
        lines.append("")
        lines.append(f"# {title}")
        lines.append("")
        lines.append("## 字段字典")
        lines.append("")
        lines.append("| # | 列 | 字段名 |")
        lines.append("|---|---|---|")
        for j, name in enumerate(colnames):
            lines.append(f"| {j + 1} | {get_column_letter(j + 1)} | {name.replace('|', chr(92) + '|')} |")
        lines.append("")
        lines.append(f"## 数据（{len(data)} 行）")
        lines.append("")
        lines.append("| " + " | ".join(c.replace("|", "\\|") for c in colnames) + " |")
        lines.append("|" + "---|" * ncol)
        for r in data:
            lines.append("| " + " | ".join(cell_md(c) for c in r) + " |")
        lines.append("")

        (out / fname).write_text("\n".join(lines), encoding="utf-8")
        index.append((fname, title, len(data), ncol))
        print(f"  ✓ {fname}  ({len(data)} 行 × {ncol} 列)")

    wb.close()

    print(f"\n生成 {len(index)} 个工作表 → {out_dir}")
    if skipped:
        print("跳过：")
        for name, why in skipped:
            print(f"  - {name!r}: {why}")


if __name__ == "__main__":
    main()
